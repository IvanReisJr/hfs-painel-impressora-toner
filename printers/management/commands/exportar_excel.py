"""Management command: python manage.py exportar_excel [--saida arquivo.xlsx]

Gera planilha Excel formatada para verificação dos níveis de toner.

Abas geradas:
  1. Impressoras  — status atual de todas as impressoras ativas
  2. Histórico    — últimas leituras dos 7 dias anteriores
  3. Sem Dados    — impressoras sem nenhuma leitura ainda

Uso:
    python manage.py exportar_excel
    python manage.py exportar_excel --saida verificacao_toner_maio.xlsx
    python manage.py exportar_excel --todas          (inclui inativas)
"""

from datetime import timedelta
from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from printers.models import Printer, TonerReading

_DEFAULT_OUTPUT = "verificacao_toner.xlsx"

# ── Palette ─────────────────────────────────────────────────────────────────
_C_HEADER_BG   = "1A3C6E"   # azul escuro
_C_HEADER_FG   = "FFFFFF"
_C_CRITICAL_BG = "F8D7DA"   # vermelho suave
_C_CRITICAL_FG = "842029"
_C_WARNING_BG  = "FFF3CD"   # amarelo suave
_C_WARNING_FG  = "664D03"
_C_OK_BG       = "D1E7DD"   # verde suave
_C_OK_FG       = "0A3622"
_C_NODATA_BG   = "E2E3E5"   # cinza
_C_NODATA_FG   = "41464B"
_C_ALT_ROW     = "F0F4FA"   # azul bem claro (linhas alternadas)
_C_CHECK_BG    = "FFF8E1"   # fundo das colunas do estagiário
_C_BORDER      = "BBCDE5"

_FONT_NAME = "Arial"

# ── Helpers ──────────────────────────────────────────────────────────────────

def _font(bold=False, size=10, color="000000"):
    return Font(name=_FONT_NAME, bold=bold, size=size, color=color)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _border() -> Border:
    side = Side(style="thin", color=_C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _font(bold=True, size=10, color=_C_HEADER_FG)
        cell.fill = _fill(_C_HEADER_BG)
        cell.alignment = _center()
        cell.border = _border()


def _toner_style(pct: int | None) -> tuple[str, str]:
    """Returns (bg_color, fg_color) for a toner percentage cell."""
    if pct is None:
        return _C_NODATA_BG, _C_NODATA_FG
    if pct <= 10:
        return _C_CRITICAL_BG, _C_CRITICAL_FG
    if pct <= 20:
        return _C_WARNING_BG, _C_WARNING_FG
    return _C_OK_BG, _C_OK_FG


def _write_cell(ws, row, col, value, bold=False, bg=None, fg="000000",
                align=None, num_fmt=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=fg)
    cell.border = _border()
    cell.alignment = align or _left()
    if bg:
        cell.fill = _fill(bg)
    if num_fmt:
        cell.number_format = num_fmt


def _pct_text(pct: int | None) -> str:
    return f"{pct}%" if pct is not None else "—"


# ── Sheet 1: Impressoras ─────────────────────────────────────────────────────

_HEADERS_IMPRESSORAS = [
    "Nome da Impressora", "Endereço IP", "Localização", "Modelo",
    "Colorida", "Protocolo",
    "Preto (K)", "Ciano (C)", "Magenta (M)", "Amarelo (Y)",
    "Status", "Última Coleta",
    "✓ Verificado", "Observação do Estagiário",
]

_COL_WIDTHS_IMPRESSORAS = [
    28, 16, 22, 22, 9, 10,
    10, 10, 10, 10,
    12, 18,
    14, 30,
]


def _build_sheet_impressoras(ws, printers) -> None:
    ws.title = "Impressoras"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # Header
    for col, (header, width) in enumerate(
        zip(_HEADERS_IMPRESSORAS, _COL_WIDTHS_IMPRESSORAS), start=1
    ):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width
    _style_header_row(ws, 1, len(_HEADERS_IMPRESSORAS))
    ws.row_dimensions[1].height = 32

    # Data rows
    for idx, printer in enumerate(printers):
        row = idx + 2
        reading = printer.latest_reading
        alt = idx % 2 == 1

        base_bg = _C_ALT_ROW if alt else "FFFFFF"

        def wc(col, value, bg=None, fg="000000", bold=False, align=None):
            _write_cell(ws, row, col, value,
                        bold=bold, bg=bg or base_bg, fg=fg, align=align)

        wc(1,  printer.name, bold=True)
        wc(2,  printer.ip_address, align=_center())
        wc(3,  printer.location or "—")
        wc(4,  printer.model_name or "—")
        wc(5,  "Sim" if printer.is_color else "Não", align=_center())
        wc(6,  printer.get_protocol_display(), align=_center())

        if reading:
            for offset, pct in enumerate([
                reading.black_pct, reading.cyan_pct,
                reading.magenta_pct, reading.yellow_pct
            ]):
                bg, fg = _toner_style(pct)
                _write_cell(ws, row, 7 + offset, _pct_text(pct),
                             bg=bg, fg=fg, align=_center())

            level = reading.alert_level
            status_map = {
                "critical": ("⚠ CRÍTICO",  _C_CRITICAL_BG, _C_CRITICAL_FG),
                "warning":  ("! ATENÇÃO",   _C_WARNING_BG,  _C_WARNING_FG),
                "ok":       ("✓ OK",        _C_OK_BG,       _C_OK_FG),
            }
            label, sbg, sfg = status_map.get(level, ("—", base_bg, "000000"))
            _write_cell(ws, row, 11, label, bg=sbg, fg=sfg,
                         bold=True, align=_center())
            _write_cell(ws, row, 12,
                         reading.collected_at.strftime("%d/%m/%Y %H:%M"),
                         bg=base_bg, align=_center())
        else:
            for col in range(7, 12):
                wc(col, "—", align=_center())
            _write_cell(ws, row, 11, "SEM DADOS",
                         bg=_C_NODATA_BG, fg=_C_NODATA_FG,
                         bold=True, align=_center())
            _write_cell(ws, row, 12, "—", bg=base_bg, align=_center())

        # Intern columns — highlighted background
        _write_cell(ws, row, 13, "", bg=_C_CHECK_BG, align=_center())
        _write_cell(ws, row, 14, "", bg=_C_CHECK_BG)

        ws.row_dimensions[row].height = 20

    # Auto-filter on header
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS_IMPRESSORAS))}1"

    # Legend below table
    legend_row = len(printers) + 4
    ws.cell(row=legend_row, column=1, value="Legenda:").font = _font(bold=True)
    for i, (label, bg, fg) in enumerate([
        ("≤ 10% — CRÍTICO",  _C_CRITICAL_BG, _C_CRITICAL_FG),
        ("≤ 20% — ATENÇÃO",  _C_WARNING_BG,  _C_WARNING_FG),
        ("> 20% — OK",        _C_OK_BG,       _C_OK_FG),
        ("Sem leitura",       _C_NODATA_BG,   _C_NODATA_FG),
    ], start=1):
        cell = ws.cell(row=legend_row + i, column=1, value=label)
        cell.font = _font(color=fg)
        cell.fill = _fill(bg)
        cell.border = _border()


# ── Sheet 2: Histórico ───────────────────────────────────────────────────────

_HEADERS_HISTORICO = [
    "Impressora", "IP", "Data/Hora Coleta",
    "Preto (K)", "Ciano (C)", "Magenta (M)", "Amarelo (Y)",
    "Protocolo", "Status",
]
_COL_WIDTHS_HISTORICO = [28, 16, 20, 10, 10, 10, 10, 10, 12]


def _build_sheet_historico(ws, readings) -> None:
    ws.title = "Histórico 7 dias"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for col, (header, width) in enumerate(
        zip(_HEADERS_HISTORICO, _COL_WIDTHS_HISTORICO), start=1
    ):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width
    _style_header_row(ws, 1, len(_HEADERS_HISTORICO))
    ws.row_dimensions[1].height = 32

    for idx, r in enumerate(readings):
        row = idx + 2
        alt = idx % 2 == 1
        base_bg = _C_ALT_ROW if alt else "FFFFFF"

        _write_cell(ws, row, 1, r.printer.name, bg=base_bg)
        _write_cell(ws, row, 2, r.printer.ip_address, bg=base_bg, align=_center())
        _write_cell(ws, row, 3,
                     r.collected_at.strftime("%d/%m/%Y %H:%M"),
                     bg=base_bg, align=_center())

        for offset, pct in enumerate([r.black_pct, r.cyan_pct, r.magenta_pct, r.yellow_pct]):
            bg, fg = _toner_style(pct)
            _write_cell(ws, row, 4 + offset, _pct_text(pct),
                         bg=bg, fg=fg, align=_center())

        _write_cell(ws, row, 8, r.protocol_used or "—", bg=base_bg, align=_center())

        level = r.alert_level
        status_map = {
            "critical": ("⚠ CRÍTICO", _C_CRITICAL_BG, _C_CRITICAL_FG),
            "warning":  ("! ATENÇÃO",  _C_WARNING_BG,  _C_WARNING_FG),
            "ok":       ("✓ OK",       _C_OK_BG,        _C_OK_FG),
        }
        label, sbg, sfg = status_map.get(level, ("—", base_bg, "000000"))
        _write_cell(ws, row, 9, label, bg=sbg, fg=sfg, bold=True, align=_center())
        ws.row_dimensions[row].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS_HISTORICO))}1"


# ── Sheet 3: Sem Dados ───────────────────────────────────────────────────────

_HEADERS_SEMDADOS = ["Nome", "IP", "Localização", "Modelo", "Protocolo", "Ativa"]
_COL_WIDTHS_SEMDADOS = [28, 16, 22, 22, 12, 8]


def _build_sheet_sem_dados(ws, printers_no_data) -> None:
    ws.title = "Sem Dados"
    ws.sheet_view.showGridLines = False

    for col, (header, width) in enumerate(
        zip(_HEADERS_SEMDADOS, _COL_WIDTHS_SEMDADOS), start=1
    ):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width
    _style_header_row(ws, 1, len(_HEADERS_SEMDADOS))
    ws.row_dimensions[1].height = 28

    for idx, p in enumerate(printers_no_data):
        row = idx + 2
        bg = _C_ALT_ROW if idx % 2 == 1 else "FFFFFF"
        _write_cell(ws, row, 1, p.name, bg=bg)
        _write_cell(ws, row, 2, p.ip_address, bg=bg, align=_center())
        _write_cell(ws, row, 3, p.location or "—", bg=bg)
        _write_cell(ws, row, 4, p.model_name or "—", bg=bg)
        _write_cell(ws, row, 5, p.get_protocol_display(), bg=bg, align=_center())
        _write_cell(ws, row, 6, "Sim" if p.is_active else "Não",
                     bg=bg, align=_center())
        ws.row_dimensions[row].height = 18


# ── Command ──────────────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = "Exporta planilha Excel formatada para verificação de toner"

    def add_arguments(self, parser):
        parser.add_argument(
            "--saida",
            type=str,
            default=_DEFAULT_OUTPUT,
            help=f"Nome do arquivo de saída (padrão: {_DEFAULT_OUTPUT})",
        )
        parser.add_argument(
            "--todas",
            action="store_true",
            default=False,
            help="Inclui impressoras inativas",
        )

    def handle(self, *args, **options):
        output_path = Path(options["saida"])
        include_all = options["todas"]

        qs = Printer.objects.all() if include_all else Printer.objects.filter(is_active=True)
        printers = list(qs.order_by("name").prefetch_related("readings"))

        since = timezone.now() - timedelta(days=7)
        readings = list(
            TonerReading.objects.filter(
                success=True, collected_at__gte=since
            ).select_related("printer").order_by("-collected_at")
        )

        printers_no_data = [p for p in printers if not p.latest_reading]
        printers_with_data = [p for p in printers if p.latest_reading]

        wb = Workbook()
        ws1 = wb.active
        _build_sheet_impressoras(ws1, printers_with_data)

        ws2 = wb.create_sheet()
        _build_sheet_historico(ws2, readings)

        if printers_no_data:
            ws3 = wb.create_sheet()
            _build_sheet_sem_dados(ws3, printers_no_data)

        wb.save(output_path)

        total = len(printers)
        critical = sum(
            1 for p in printers_with_data
            if p.latest_reading and p.latest_reading.alert_level == "critical"
        )
        warning = sum(
            1 for p in printers_with_data
            if p.latest_reading and p.latest_reading.alert_level == "warning"
        )

        self.stdout.write(self.style.SUCCESS(f"\nPlanilha gerada: {output_path.resolve()}"))
        self.stdout.write(f"  Total de impressoras : {total}")
        self.stdout.write(self.style.ERROR(f"  Critico  (<=10%) : {critical}") if critical else f"  Critico  (<=10%) : {critical}")
        self.stdout.write(self.style.WARNING(f"  Atencao  (<=20%) : {warning}") if warning else f"  Atencao  (<=20%) : {warning}")
        self.stdout.write(f"  Sem dados            : {len(printers_no_data)}")
        self.stdout.write(f"  Leituras (7 dias)    : {len(readings)}\n")
        self.stdout.write("  Abra o arquivo no Excel e entregue ao estagiario.")
        self.stdout.write("  Colunas 'Verificado' e 'Observacao' estao destacadas em amarelo.\n")
